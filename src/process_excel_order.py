import os
import shutil
import logging
import pandas as pd
import re
from datetime import datetime
from openpyxl import load_workbook
from database import get_db_engine

# --- 設定 Log 紀錄 ---
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler("process_excel.log", encoding='utf-8'),
        logging.StreamHandler()
    ]
)

# --- 路徑設定 ---
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EXCEL_SOURCE_DIR = os.path.join(BASE_DIR, 'uploads', 'daily_excel')
EXCEL_PROCESSED_DIR = os.path.join(BASE_DIR, 'uploads', 'daily_excel', 'processed')

def get_mawb_no(filepath, filename):
    """
    取得主提單號 (MAWB)
    邏輯:
    1. 嘗試從檔案 A1 儲存格讀取
    2. 驗證 A1 內容:
       - 若為 "數字+英文字母" (Alphanumeric) -> 視為有效 MAWB
       - 若包含中文、特殊符號或為空 -> 視為無效 (可能是標題)，改抓取檔名
    3. 若 A1 無效，則使用檔名 (去除副檔名)
    """
    mawb_no = None
    a1_value = None
    
    # --- 步驟 1: 讀取 A1 儲存格 ---
    try:
        if filename.lower().endswith('.csv'):
            # CSV: 只讀取第1行第1列
            df_temp = pd.read_csv(filepath, nrows=1, header=None, usecols=[0])
            if not df_temp.empty:
                val = str(df_temp.iloc[0, 0]).strip()
                if val and val.lower() not in ['nan', 'none', '']:
                    a1_value = val
        else:
            # Excel: 使用 openpyxl
            wb = load_workbook(filepath, read_only=True, data_only=True)
            ws = wb.active
            cell_value = ws['A1'].value
            wb.close()
            
            if cell_value:
                val = str(cell_value).strip()
                if val and val.lower() not in ['nan', 'none', '']:
                    a1_value = val
    except Exception as e:
        logging.warning(f"讀取 A1 欄位時發生錯誤: {e}")

    # --- 步驟 2: 驗證 A1 內容 (智慧判斷) ---
    is_valid_a1 = False
    if a1_value:
        # Regex: ^[A-Za-z0-9]+$ 表示字串從頭到尾只能由 英文字母(大小寫) 或 數字 組成
        if re.match(r'^[A-Za-z0-9]+$', a1_value):
            mawb_no = a1_value
            is_valid_a1 = True
            logging.info(f"A1 欄位驗證通過 (數字+英文): {mawb_no}")
        else:
            logging.info(f"A1 欄位內容 '{a1_value}' 不符合純英數格式 (可能是中文標題)，將忽略 A1。")

    # --- 步驟 3: 若 A1 無效，使用檔名 ---
    if not is_valid_a1:
        # 去除副檔名 (例如: 25040104EX.xlsx -> 25040104EX)
        mawb_no = os.path.splitext(filename)[0].strip()
        logging.info(f"使用檔名作為主提單號 (MAWB): {mawb_no}")

    return mawb_no

def process_data_old_format(df):
    """
    處理舊格式邏輯 (Header在第4列)
    依賴欄位名稱: 分提單號碼, 貨物編號, 货物名称...
    """
    # 欄位正規化
    df.columns = [str(c).strip().replace('\n', '') for c in df.columns]
    
    required_cols = ['分提單號碼', '貨物編號', '货物名称', '數量', '單價金額', '發票總金額']
    missing_cols = [c for c in required_cols if c not in df.columns]
    if missing_cols:
        return None, f"舊格式欄位缺失: {missing_cols}"

    # 清洗與提取
    df['分提單號碼'] = df['分提單號碼'].ffill() # 填補合併儲存格
    df = df[df['分提單號碼'].notna() & df['貨物編號'].notna()].copy()
    
    db_df = pd.DataFrame()
    db_df['hawb_no'] = df['分提單號碼'].astype(str).str.strip()
    db_df['item_no'] = pd.to_numeric(df['貨物編號'], errors='coerce').fillna(0).astype(int)
    db_df['description_original'] = df['货物名称'].astype(str).str.strip()
    db_df['qty'] = pd.to_numeric(df['數量'], errors='coerce').fillna(0)
    
    if '數量單位' in df.columns:
        db_df['qty_unit'] = df['數量單位'].astype(str).str.strip()
    if '淨重' in df.columns:
        db_df['net_weight'] = pd.to_numeric(df['淨重'], errors='coerce').fillna(0)
        
    db_df['unit_price'] = pd.to_numeric(df['單價金額'], errors='coerce').fillna(0)
    db_df['total_amount'] = pd.to_numeric(df['發票總金額'], errors='coerce').fillna(0)
    db_df['currency'] = 'TWD'
    
    # 關係人
    if '進口人英文名稱' in df.columns:
        db_df['consignee_name'] = df['進口人英文名稱'].astype(str).str.strip()
    if '進口人統一編號' in df.columns:
        db_df['consignee_id'] = df['進口人統一編號'].astype(str).str.strip()
    if '進口人電話' in df.columns:
        db_df['consignee_phone'] = df['進口人電話'].astype(str).str.strip()
        
    return db_df, "Old Format"

def process_data_new_format(df):
    """
    處理新格式邏輯 (Header在第3列)
    依賴固定欄位索引: A(0), D(3), J(9), K(10), N(13), O(14)
    """
    # 檢查欄位數量是否足夠 (至少要有 O 欄, index 14)
    if len(df.columns) < 15:
        return None, "新格式欄位數量不足 (至少需15欄)"

    # 提取關鍵欄位 (使用 iloc 依索引取值)
    # A=0 (分提單), D=3 (品名), J=9 (數量), K=10 (單位), N=13 (單價), O=14 (總價)
    try:
        # 建立臨時 DataFrame 處理
        temp_df = pd.DataFrame()
        temp_df['hawb_raw'] = df.iloc[:, 0]  # A欄
        temp_df['desc_raw'] = df.iloc[:, 3]  # D欄
        temp_df['qty_raw'] = df.iloc[:, 9]   # J欄
        temp_df['unit_raw'] = df.iloc[:, 10] # K欄
        temp_df['price_raw'] = df.iloc[:, 13] # N欄
        temp_df['total_raw'] = df.iloc[:, 14] # O欄
    except IndexError:
        return None, "新格式欄位索引讀取失敗"

    # 處理拆單 (Merged Cells) - A欄分提單號向下填補
    temp_df['hawb_raw'] = temp_df['hawb_raw'].ffill()

    # 過濾有效行：分提單號與品名都必須有值
    temp_df = temp_df[temp_df['hawb_raw'].notna() & temp_df['desc_raw'].notna()].copy()
    
    if temp_df.empty:
        return None, "新格式無有效資料"

    # 自動產生項次 (Item No)
    # 邏輯: 針對同一個分提單號，依序編號 1, 2, 3...
    temp_df['item_no'] = temp_df.groupby('hawb_raw').cumcount() + 1

    # 映射到資料庫格式
    db_df = pd.DataFrame()
    db_df['hawb_no'] = temp_df['hawb_raw'].astype(str).str.strip()
    db_df['item_no'] = temp_df['item_no']
    db_df['description_original'] = temp_df['desc_raw'].astype(str).str.strip()
    db_df['qty'] = pd.to_numeric(temp_df['qty_raw'], errors='coerce').fillna(0)
    db_df['qty_unit'] = temp_df['unit_raw'].astype(str).str.strip()
    db_df['unit_price'] = pd.to_numeric(temp_df['price_raw'], errors='coerce').fillna(0)
    db_df['total_amount'] = pd.to_numeric(temp_df['total_raw'], errors='coerce').fillna(0)
    db_df['currency'] = 'TWD'
    
    # 新格式目前沒有指定關係人欄位，先留空
    db_df['consignee_name'] = None
    db_df['consignee_id'] = None
    db_df['consignee_phone'] = None

    return db_df, "New Format"

def process_excel_file(filepath, filename):
    engine = get_db_engine()
    if not engine: return

    try:
        # 1. 取得主提單號 (優先 A1 且需為英數，失敗則用檔名)
        mawb_no = get_mawb_no(filepath, filename)
        
        if not mawb_no:
            logging.error(f"錯誤: 檔案 {filename} 無法取得有效主提單號，跳過。")
            return

        logging.info(f"開始處理檔案: {filename}, MAWB: {mawb_no}")
        
        # 2. 格式偵測與讀取
        db_df = None
        format_type = "Unknown"
        error_msg = ""

        # [策略 A] 先嘗試「舊格式」 (Header=3, Row 4)
        try:
            if filename.lower().endswith('.csv'):
                df_old = pd.read_csv(filepath, header=3)
            else:
                df_old = pd.read_excel(filepath, header=3)
            
            # 檢查舊格式關鍵欄位是否存在
            if '分提單號碼' in [str(c).strip().replace('\n', '') for c in df_old.columns]:
                db_df, format_type = process_data_old_format(df_old)
                if db_df is None: error_msg = format_type
            else:
                error_msg = "非舊格式特徵"
        except Exception as e:
            error_msg = f"舊格式讀取錯誤: {e}"

        # [策略 B] 若舊格式失敗，嘗試「新格式」 (Header=2, Row 3)
        if db_df is None:
            logging.info(f"舊格式匹配失敗 ({error_msg})，嘗試新格式...")
            try:
                if filename.lower().endswith('.csv'):
                    df_new = pd.read_csv(filepath, header=2)
                else:
                    df_new = pd.read_excel(filepath, header=2)
                
                # 直接進入新格式處理邏輯
                db_df, format_type = process_data_new_format(df_new)
                if db_df is None:
                    logging.error(f"新格式匹配失敗: {format_type}")
                    return
            except Exception as e:
                logging.error(f"新格式讀取錯誤: {e}")
                return

        # 3. 寫入資料庫
        if db_df is not None and not db_df.empty:
            db_df['mawb_no'] = mawb_no
            db_df['processing_status'] = 'PENDING'
            
            logging.info(f"偵測為 [{format_type}]，解析出 {len(db_df)} 筆資料。")
            db_df.to_sql('table_a_raw', con=engine, if_exists='append', index=False)
            logging.info(f"成功寫入資料庫！")

            # 4. 移動檔案
            if not os.path.exists(EXCEL_PROCESSED_DIR):
                os.makedirs(EXCEL_PROCESSED_DIR)
            shutil.move(filepath, os.path.join(EXCEL_PROCESSED_DIR, filename))
            logging.info(f"檔案已移至 processed 資料夾")
        else:
            logging.warning("無法解析出有效資料，跳過匯入。")

    except Exception as e:
        logging.error(f"處理檔案 {filename} 時發生未預期錯誤: {e}", exc_info=True)

def main():
    if not os.path.exists(EXCEL_SOURCE_DIR):
        os.makedirs(EXCEL_SOURCE_DIR)
        return

    files = [f for f in os.listdir(EXCEL_SOURCE_DIR) if f.lower().endswith(('.xlsx', '.xls', '.csv'))]
    if not files:
        logging.info("目前沒有需要處理的 Excel 檔案。")
        return

    logging.info(f"掃描到 {len(files)} 個檔案，開始作業...")
    for filename in files:
        filepath = os.path.join(EXCEL_SOURCE_DIR, filename)
        process_excel_file(filepath, filename)

if __name__ == "__main__":
    main()